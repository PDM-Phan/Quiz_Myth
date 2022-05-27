import openpyxl as op
from random import randint
from time import sleep
from Functions import verificar_resposta, guarda_perguntas, printa_pergunta
from Functions import Ranking_Functions, Menu_Quiz, Functions_Admin

# Main Program

# Criando ambientes de trabalho futuros
quiz_book = op.load_workbook('Pasta1.xlsx')
sheet_quiz = quiz_book['Quiz']
sheet_cadastro = quiz_book['Cadastro']
sheet_admin = quiz_book['Admin']
sheet_rankings = quiz_book['Ranking']

# Debug linhas perguntas
# linhas_totais = sheet_quiz.max_row
# print(linhas_totais)

# coloca os ranks em uma variavel lista
ranking = Ranking_Functions.carregando_ranking(sheet_rankings)

# Criando/logando jogador
jogador_admin = Menu_Quiz.Janela_principal(quiz_book, sheet_cadastro, sheet_admin)

# Determina o maximo de perguntas que existe na planilha
linhas = sheet_quiz.max_row - 1
# Recebe todas as perguntas da planilha
perguntas, alternativas = guarda_perguntas(sheet_quiz, linhas)
Perguntas_Disponiveis = []

# determina as funçoes para cada usuario, se é admin ou se é jogador.
if len(jogador_admin) > 1:
    while len(Perguntas_Disponiveis) != linhas:
        # Apresentando o quiz
        Menu_Quiz.Menu("QUIZ SERES MITOLOGICOS", jogador_admin[0])
        print('\033[1;4mOBS:\033[m OS PONTOS SÃO CAUCULADOS BASEADOS NA QUANTIDADE DE '
              '\033[1mPERGUNTAS_RESPONDIDAS_TOTAIS\033[m COM AS \033[1mQUANTIDADES DE PERGUNTAS_CORRETAMENTE\033[m.\n ')
        pontos = perguntas_respondidas = 0
        # Loop onde o jogo so termina se o jogador quiser, ou não tiver mais nenhuma pergunta disponivel

        while len(Perguntas_Disponiveis) != linhas:
            # Escolhe uma pergunta "aleatoriamente"
            num_perg = randint(0, linhas - 1)
            # Se a pergunta não ja tiver sido escolhida, o programa procede com o jogo
            if num_perg not in Perguntas_Disponiveis:
                Perguntas_Disponiveis.append(num_perg)
                printa_pergunta(perguntas, num_perg)
                while True:
                    resposta = input('>> Qual a alternativa correta? ').upper().strip()
                    # Verificando se a resposta do usuario é válida
                    if resposta not in 'ABC' or resposta == '':
                        print('\033[31m!! Digite uma alternativa válida !!\033[m')
                        continue
                    else:
                        # Verifica a resposta do usuario é correda ou não
                        verificador = verificar_resposta(alternativas, num_perg, resposta, sheet_quiz)
                        if verificador > 0:
                            if verificador == 2:
                                pontos += 2
                                perguntas_respondidas += 1
                            else:
                                pontos += 1
                                perguntas_respondidas += 1
                            print(f'\033[1;32m!!VOCE ACERTOU!!\033[m A dificuldade da questão que você respondeu foi '
                                  f'{"Dificil" if verificador == 2 else "Fácil"}.')
                        else:
                            print(f'\033[1;31m!!VOCÊ ERROU!!\033[m A reposta certa era a alternativa '
                                  f'\033[1;4m{alternativas[num_perg]}\033[m.')
                            perguntas_respondidas += 1
                        break
            # Se ja foi escolhida, então será gerada outra pergunta
            else:
                continue

            # Número minimo de perguntas, antes do usuario ter a opção de parar o jogo
            if len(Perguntas_Disponiveis) >= 5:
                # Após a primeira onda de perguntas, o programa irá perguntar se o jogador irá querer continuar.
                loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()
                while loop_continue not in 'SN' or loop_continue == '':
                    print('\033[31m!! Digite uma alternativa válida !!\033[m')
                    loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()

                # Caso o usuario não queira, o programa irá salvar os resultados da partida e então irá parar
                if loop_continue == 'N':
                    print('Finalizando o QUIZ...')
                    sleep(1)
                    print('Registrando jogador...')
                    # caucula os pontos ganhos na partida, e após isso adciona no perfil do usuario
                    rank = Ranking_Functions.caucular_rank(pontos, perguntas_respondidas)
                    jogador_admin[1] += rank
                    print(f'\033[1;32mParabéns!! Vcoê conseguiu \033[4m{rank}\033[m \033[1;32mpontos!!\033[m')
                    # atualiza o ranking da variavel, colocando o usuario na sua posiçao
                    Ranking_Functions.alterando_ranking(ranking, jogador_admin)
                    # salva esse novo ranking na planilha
                    Ranking_Functions.alterando_ranking_planilha(sheet_rankings, ranking, quiz_book)
                    Ranking_Functions.printando_ranking(ranking)
                    sleep(1)
                    print('\nQUIZ FINALIZADO.')
                    break

                else:
                    print('Criando uma nova pergunta...')
                    sleep(1)

            else:
                print('Criando uma nova pergunta...')
                sleep(1)
        jogador_admin.pop(0)
        jogador_admin[1] = 0

else:
    while True:
        Functions_Admin.menu_admin(jogador_admin[0])
        try:
            controle = int(input('>> '))
            if 1 > controle > 4:
                raise ValueError

            if controle == 1:
                Functions_Admin.adcionar_pergunta(sheet_quiz, quiz_book)

            elif controle == 2:
                Functions_Admin.remover_pergunta(sheet_quiz, quiz_book)

            elif controle == 3:
                Functions_Admin.mostrar_perguntas_admin(perguntas, sheet_quiz)
                1
            elif controle > 3:
                break

        except ValueError:
            print('[ERRO] Digite uma opção válida.\n'
                  'tente novamente.')
            continue

